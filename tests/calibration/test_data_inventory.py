from __future__ import annotations

import csv
from pathlib import Path

import pytest

from pybamm_w10.calibration.data import (
    CAPACITY_COLUMNS,
    CYCLING_COLUMNS,
    DatasetValidationError,
    W10_DIAGNOSTIC_NODES,
    build_diagnostic_inventory,
    load_cycle0_capacity_curve,
    write_diagnostic_inventory,
)


def _write_capacity(path: Path, index: int, endpoint: float) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(CAPACITY_COLUMNS)
        writer.writerow((index, "W10", 0.0, 0.0, 4.2, 0.0))
        writer.writerow((index, "W10", 1.0, -0.24, 2.5, endpoint))


def _write_readme(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    first = workbook.active
    first.title = "Capacity_test"
    for sheet_name in ("HPPC", "EIS"):
        workbook.create_sheet(sheet_name)
    for worksheet in workbook.worksheets:
        worksheet.cell(1, 8).value = "W10"
        for row, node in enumerate(W10_DIAGNOSTIC_NODES, start=2):
            worksheet.cell(row, 8).value = node
    workbook.save(path)


def make_w10_data_root(root: Path) -> Path:
    m50t = root / "LG M50T"
    diagnostics = m50t / "_processed_mat"
    for index in range(1, 16):
        _write_capacity(diagnostics / f"W10_capacity_diagnostic_{index:02d}.csv", index, 4.9 - index / 100)
    mat_root = m50t / "cycling" / "W10"
    csv_root = m50t / "cycling" / "w10_dataset"
    mat_root.mkdir(parents=True)
    csv_root.mkdir(parents=True)
    for index in range(1, 15):
        (mat_root / f"W10-{index}.mat").write_bytes(b"MAT")
        with (csv_root / f"W10-{index}.csv").open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(CYCLING_COLUMNS)
            writer.writerow((0, 0, 3.2, 0, 0, 0))
    _write_readme(m50t / "README.xlsx")
    return root


def test_inventory_validates_w10_files_and_allows_stage1_without_w10_hppc_eis(workspace_tmp) -> None:
    root = make_w10_data_root(workspace_tmp / "data")
    inventory = build_diagnostic_inventory(root)

    assert inventory["diagnostic_nodes"] == list(W10_DIAGNOSTIC_NODES)
    assert len(inventory["capacity_diagnostics"]) == 15
    assert len(inventory["cycling_mat"]) == len(inventory["cycling_csv"]) == 14
    assert inventory["capacity_diagnostics"][0]["endpoint_capacity_ah"] == pytest.approx(4.89)
    assert inventory["capacity_diagnostics"][-1]["endpoint_capacity_ah"] is None
    assert inventory["aging_calibration_gate"] == {
        "status": "AGING_CALIBRATION_READY",
        "reason": None,
        "w10_hppc_files": [],
        "w10_eis_files": [],
    }
    path = write_diagnostic_inventory(workspace_tmp / "diagnostic_inventory.json", inventory)
    assert path.is_file()


def test_inventory_rejects_non_monotonic_capacity(workspace_tmp) -> None:
    root = make_w10_data_root(workspace_tmp / "data")
    path = root / "LG M50T" / "_processed_mat" / "W10_capacity_diagnostic_03.csv"
    with path.open("a", newline="", encoding="utf-8") as handle:
        csv.writer(handle).writerow((3, "W10", 2.0, -0.24, 2.5, 0.1))
    with pytest.raises(DatasetValidationError, match="capacity is not monotonic"):
        build_diagnostic_inventory(root)


def test_cycle0_curve_loader_uses_zero_padded_first_diagnostic_only(workspace_tmp) -> None:
    root = make_w10_data_root(workspace_tmp / "data")
    capacity, voltage = load_cycle0_capacity_curve(root)

    assert capacity.tolist() == pytest.approx([0.0, 4.89])
    assert voltage.tolist() == pytest.approx([4.2, 2.5])
