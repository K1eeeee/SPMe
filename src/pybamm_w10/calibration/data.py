"""Read-only W10 calibration-data inventory and validation."""

from __future__ import annotations

import csv
from hashlib import sha256
import math
from pathlib import Path
from typing import Any, Iterable

import numpy as np

from .artifacts import write_calibration_json


W10_DIAGNOSTIC_NODES = (0, 25, 75, 122, 146, 148, 151, 159, 188, 225, 250, 275, 300, 325, 350)
ANCHOR_NODES = (0,)
CALIBRATION_NODES = (25, 75, 122, 146, 148, 151, 159, 188)
PUBLIC_CALIBRATION_NODES = ANCHOR_NODES + CALIBRATION_NODES
CAPACITY_COLUMNS = ("diagnostic_number", "cell_id", "time_s", "current_a", "voltage_v", "capacity_ah")
CYCLING_COLUMNS = (
    "I_full_vec_M1_NMC25degC",
    "Step_Index_full_vec_M1_NMC25degC",
    "V_full_vec_M1_NMC25degC",
    "ch_cap_full_vec_M1_NMC25degC",
    "dis_cap_full_vec_M1_NMC25degC",
    "t_full_vec_M1_NMC25degC",
)


class DatasetValidationError(ValueError):
    """A raw input failed an immutable inventory invariant."""


def _numeric_suffix(path: Path) -> int:
    try:
        return int(path.stem.rsplit("_", 1)[-1].rsplit("-", 1)[-1])
    except ValueError as exc:
        raise DatasetValidationError(f"unrecognised W10 file name: {path.name}") from exc


def _sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        while block := handle.read(1024 * 1024):
            digest.update(block)
    return digest.hexdigest()


def _inventory_file(path: Path, root: Path) -> dict[str, object]:
    return {
        "relative_path": path.relative_to(root).as_posix(),
        "size_bytes": path.stat().st_size,
        "sha256": _sha256(path),
    }


def _require_exact_files(files: Iterable[Path], count: int, kind: str) -> list[Path]:
    result = sorted(files, key=_numeric_suffix)
    if len(result) != count:
        raise DatasetValidationError(f"expected {count} {kind} files, found {len(result)}")
    suffixes = [_numeric_suffix(path) for path in result]
    if suffixes != list(range(1, count + 1)):
        raise DatasetValidationError(f"{kind} file indices must be exactly 1..{count}: {suffixes}")
    return result


def _capacity_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            header = tuple(next(reader))
        except StopIteration as exc:
            raise DatasetValidationError(f"empty capacity diagnostic: {path.name}") from exc
        if header != CAPACITY_COLUMNS:
            raise DatasetValidationError(f"unexpected capacity header in {path.name}: {header}")
        rows = [row for row in reader if tuple(row) != header]
    if not rows:
        raise DatasetValidationError(f"capacity diagnostic has no data rows: {path.name}")
    if any(len(row) != len(CAPACITY_COLUMNS) for row in rows):
        raise DatasetValidationError(f"malformed capacity diagnostic row in {path.name}")
    return [dict(zip(CAPACITY_COLUMNS, row, strict=True)) for row in rows]


def _read_capacity_endpoint(path: Path) -> float:
    """Read a validated capacity end point; raw files are never altered."""
    rows = _capacity_rows(path)
    times: list[float] = []
    capacities: list[float] = []
    for row in rows:
        if row["cell_id"] != "W10":
            raise DatasetValidationError(f"capacity diagnostic is not W10: {path.name}")
        try:
            values = [float(row[name]) for name in ("time_s", "current_a", "voltage_v", "capacity_ah")]
        except ValueError as exc:
            raise DatasetValidationError(f"non-numeric capacity row in {path.name}") from exc
        if not all(math.isfinite(value) for value in values):
            raise DatasetValidationError(f"non-finite capacity row in {path.name}")
        times.append(values[0])
        capacities.append(values[-1])
    if any(second < first for first, second in zip(times, times[1:])):
        raise DatasetValidationError(f"capacity time is not monotonic: {path.name}")
    if any(second < first for first, second in zip(capacities, capacities[1:])):
        raise DatasetValidationError(f"capacity is not monotonic: {path.name}")
    return capacities[-1]


def load_cycle0_capacity_curve(data_root: Path) -> tuple[np.ndarray, np.ndarray]:
    """Return only the permitted cycle-0 capacity/voltage curve.

    This deliberately names the first diagnostic file directly.  It cannot
    enumerate, parse, or expose any cycle-250--350 holdout target.
    """
    path = data_root.resolve() / "LG M50T" / "_processed_mat" / "W10_capacity_diagnostic_01.csv"
    rows = _capacity_rows(path)
    capacities: list[float] = []
    voltages: list[float] = []
    for row in rows:
        if row["cell_id"] != "W10":
            raise DatasetValidationError(f"capacity diagnostic is not W10: {path.name}")
        try:
            capacity = float(row["capacity_ah"])
            voltage = float(row["voltage_v"])
        except ValueError as exc:
            raise DatasetValidationError(f"non-numeric cycle-0 capacity row in {path.name}") from exc
        if not math.isfinite(capacity) or not math.isfinite(voltage):
            raise DatasetValidationError(f"non-finite cycle-0 capacity row in {path.name}")
        capacities.append(capacity)
        voltages.append(voltage)
    if len(capacities) < 2 or any(second < first for first, second in zip(capacities, capacities[1:])):
        raise DatasetValidationError(f"cycle-0 capacity is not monotonic: {path.name}")
    return np.asarray(capacities, dtype=float), np.asarray(voltages, dtype=float)


def _validate_cycling_csv_header(path: Path) -> None:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            header = tuple(next(reader))
            first_data = next(reader)
        except StopIteration as exc:
            raise DatasetValidationError(f"cycling CSV is incomplete: {path.name}") from exc
    if header != CYCLING_COLUMNS:
        raise DatasetValidationError(f"unexpected cycling header in {path.name}: {header}")
    try:
        values = [float(value) for value in first_data]
    except ValueError as exc:
        raise DatasetValidationError(f"non-numeric first cycling row in {path.name}") from exc
    if len(values) != len(CYCLING_COLUMNS) or not all(math.isfinite(value) for value in values):
        raise DatasetValidationError(f"non-finite first cycling row in {path.name}")


def _readme_w10_nodes(path: Path) -> dict[str, tuple[int, ...]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is part of the declared environment
        raise DatasetValidationError("openpyxl is required to verify README.xlsx") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, tuple[int, ...]] = {}
    for sheet in ("Capacity_test", "HPPC", "EIS"):
        if sheet not in workbook.sheetnames:
            raise DatasetValidationError(f"README.xlsx is missing {sheet} sheet")
        worksheet = workbook[sheet]
        if worksheet.cell(1, 8).value != "W10":
            raise DatasetValidationError(f"README.xlsx {sheet} sheet has no W10 column")
        values = tuple(
            int(value)
            for value in (worksheet.cell(row, 8).value for row in range(2, worksheet.max_row + 1))
            if isinstance(value, (int, float))
        )
        result[sheet] = values
    if any(result[name] != W10_DIAGNOSTIC_NODES for name in result):
        raise DatasetValidationError("README.xlsx W10 diagnostic nodes do not match the fixed schedule")
    return result


def build_diagnostic_inventory(data_root: Path) -> dict[str, object]:
    """Validate and hash immutable W10 calibration inputs without solving."""
    root = data_root.resolve()
    capacity_root = root / "LG M50T" / "_processed_mat"
    cycling_mat_root = root / "LG M50T" / "cycling" / "W10"
    cycling_csv_root = root / "LG M50T" / "cycling" / "w10_dataset"
    readme_path = root / "LG M50T" / "README.xlsx"
    for path in (capacity_root, cycling_mat_root, cycling_csv_root, readme_path):
        if not path.exists():
            raise DatasetValidationError(f"missing required W10 input: {path}")

    capacity_files = _require_exact_files(
        capacity_root.glob("W10_capacity_diagnostic_*.csv"), len(W10_DIAGNOSTIC_NODES), "capacity diagnostic"
    )
    cycling_mat_files = _require_exact_files(cycling_mat_root.glob("W10-*.mat"), 14, "cycling MAT")
    cycling_csv_files = _require_exact_files(cycling_csv_root.glob("W10-*.csv"), 14, "cycling CSV")
    endpoints = [_read_capacity_endpoint(path) for path in capacity_files]
    for path in cycling_csv_files:
        _validate_cycling_csv_header(path)
    readme_nodes = _readme_w10_nodes(readme_path)

    # Stage 1 uses only capacity/cycling data.  HPPC/EIS belongs to stage 2.
    hppc = sorted(capacity_root.glob("W10*HPPC*"))
    eis = sorted(capacity_root.glob("W10*EIS*"))
    return {
        "inventory_schema_version": 1,
        "cell_id": "W10",
        "diagnostic_nodes": list(W10_DIAGNOSTIC_NODES),
        "readme_nodes": {name: list(values) for name, values in readme_nodes.items()},
        "capacity_diagnostics": [
            {
                **_inventory_file(path, root),
                "cycle": node,
                # Holdout values are validated but never exposed through the
                # ordinary inventory consumed by calibration code.
                "endpoint_capacity_ah": endpoint if node in PUBLIC_CALIBRATION_NODES else None,
            }
            for path, node, endpoint in zip(capacity_files, W10_DIAGNOSTIC_NODES, endpoints, strict=True)
        ],
        "cycling_mat": [_inventory_file(path, root) for path in cycling_mat_files],
        "cycling_csv": [_inventory_file(path, root) for path in cycling_csv_files],
        "aging_calibration_gate": {
            "status": "AGING_CALIBRATION_READY",
            "reason": None,
            "w10_hppc_files": [_inventory_file(path, root) for path in hppc],
            "w10_eis_files": [_inventory_file(path, root) for path in eis],
        },
    }


def write_diagnostic_inventory(path: Path, inventory: dict[str, object]) -> Path:
    """Atomically persist an already validated inventory under the writable workspace."""
    return write_calibration_json(path, inventory)
